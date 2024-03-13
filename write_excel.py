import openpyxl

def find_empty_row(sheet):
    # 寻找第一个空行的行号
    row_num = 1
    while sheet.cell(row=row_num, column=1).value is not None:
        row_num += 1
    return row_num

def write_data_to_excel(excel_file, col2_data, col3_data):
    # 打开工作簿
    workbook = openpyxl.load_workbook(excel_file)

    # 选择默认的工作表
    sheet = workbook.active

    # 寻找第一个空行
    empty_row = find_empty_row(sheet)

    # 写入行号
    sheet.cell(row=empty_row, column=1, value=empty_row-1)

    # 接受的两个参数分别写入第二列和第三列
    sheet.cell(row=empty_row, column=2, value=col2_data)
    sheet.cell(row=empty_row, column=3, value=col3_data)

    # 保存工作簿到文件
    workbook.save(excel_file)


def write_list_to_excel(excel_file, data_list):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    empty_row = find_empty_row(sheet)

    sheet.cell(row=empty_row, column=1, value=empty_row-1)
    for index, data in enumerate(data_list, start=2):
        sheet.cell(row=empty_row, column=index, value=data)

    workbook.save(excel_file)
if __name__ == "__main__":
    # 示例用法
    excel_file = "experimental_data/car.xlsx"
    col2_data = "0.2"
    col3_data = "0.3"

    # write_data_to_excel(excel_file, col2_data, col3_data)
    data_list = ["0.4", "0.5", "0.6"]
    write_list_to_excel("experimental_data/ball.xlsx", data_list)
